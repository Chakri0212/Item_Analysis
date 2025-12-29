import pandas as pd
import numpy as np

import getpass as gt
import psycopg2
from sqlalchemy import create_engine, sql,event

import os
from openpyxl import load_workbook

import re
import datetime
import warnings
import sys

import matplotlib.pyplot as plt
import seaborn as sns
from scipy import stats

class color:
   PURPLE = '\033[95m'
   CYAN = '\033[96m'
   DARKCYAN = '\033[36m'
   BLUE = '\033[94m'
   GREEN = '\033[92m'
   YELLOW = '\033[93m'
   RED = '\033[91m'
   BOLD = '\033[1m'
   UNDERLINE = '\033[4m'
   END = '\033[0m'

def db_con():
    #make a connection to the database with your credentials
    '''
    user = gt.getpass('Enter db username : ')
    pswd = gt.getpass('Enter db password : ')
    db = input('Enter database : ') '''
    
    dbname = 'redshiftapps'
    user = 'reporting_user'
    pword = 'User_4_Reporting'
    host = 'redshift-apps-clusterredshift-19qcp828fizxm.ctebqc6bt0fq.us-east-1.redshift.amazonaws.com'
    port = '5439'
    
    '''
    try:
        engine

    except Exception:
        engine = create_engine('postgresql://' + user + ':' + pword + '@' + host + ':' + port + '/' + dbname)
    # Patch: Remove 'standard_conforming_strings' for Redshift
    @event.listens_for(engine, "connect")
    def do_connect(dbapi_connection, connection_record):
        try:
            dbapi_connection.cursor().execute("SET standard_conforming_strings TO off")
        except Exception:
            pass  # Ignore if not supported 
        
        '''

    engine = create_engine('postgresql://' + user + ':' + pword + '@' + host + ':' + port + '/' + dbname) 

    '''
    # Test the engine connection
    try:
        with engine.connect() as conn:
            print("Database connection successful.")
    except Exception as e:
        print(f"Database connection failed: {e}")
        return '''

    return engine  

    

def teleg_msg(msg):
    import telegram_send

    telegram_token = '1361504293:AAGg_ssdc4XNDLYRwv0dZF8Yqvih6r2ah6I'
    chat_id = '-1001495643344'
    path_config = telegram_send.get_config_path()

    with open(path_config, 'w') as f:
      f.write(f'[telegram]\ntoken = {telegram_token}\nchat_id = {chat_id}')

    telegram_send.send(messages = [msg])

def excel_write(framesDict, resultsPath, fileName, index = False):
    #initialze the excel writer
    writer = pd.ExcelWriter(resultsPath+fileName, engine='xlsxwriter')
    
    #now loop thru and put each on a specific sheet
    for sheet, frame in  framesDict.items(): # .use .items for python 3.X
        frame.to_excel(writer, sheet_name = sheet, index = index)
    
    #critical last step
    writer.save()

def toCamelCase(str_lst, sep = '_'):
    snake_lst = []
    for token in str_lst:
        components = token.split(sep)
        snakeName =  components[0] + ''.join(x.title() for x in components[1:])
        snake_lst.append(snakeName)
    return snake_lst
    
def append_df_to_excel(filename, df, sheet_name='Sheet1', startrow=None, bufferrow = 0,
                       truncate_sheet=False, sheet_visible = True,
                       **to_excel_kwargs):
    """
    Append a DataFrame [df] to existing Excel file [filename]
    into [sheet_name] Sheet.
    If [filename] doesn't exist, then this function will create it.

    @param filename: File path or existing ExcelWriter
                     (Example: '/path/to/file.xlsx')
    @param df: DataFrame to save to workbook
    @param sheet_name: Name of sheet which will contain DataFrame.
                       (default: 'Sheet1')
    @param startrow: upper left cell row to dump data frame.
                     Per default (startrow=None) calculate the last row
                     in the existing DF and write to the next row...
    @param truncate_sheet: truncate (remove and recreate) [sheet_name]
                           before writing DataFrame to Excel file
    @param to_excel_kwargs: arguments which will be passed to `DataFrame.to_excel()`
                            [can be a dictionary]
    @return: None

    Usage examples:

    >>> append_df_to_excel('d:/temp/test.xlsx', df)

    >>> append_df_to_excel('d:/temp/test.xlsx', df, header=None, index=False)

    >>> append_df_to_excel('d:/temp/test.xlsx', df, sheet_name='Sheet2',
                           index=False)

    >>> append_df_to_excel('d:/temp/test.xlsx', df, sheet_name='Sheet2', 
                           index=False, startrow=25)

    (c) [MaxU](https://stackoverflow.com/users/5741205/maxu?tab=profile)
    """
    # Excel file doesn't exist - saving and exiting
    if not os.path.isfile(filename):
        df.to_excel(
            filename,
            sheet_name=sheet_name, 
            startrow=startrow if startrow is not None else 0, 
            **to_excel_kwargs)
        return
    
    # ignore [engine] parameter if it was passed
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')

    writer = pd.ExcelWriter(filename, engine='openpyxl', mode='a')

    # try to open an existing workbook
    writer.book = load_workbook(filename)
    
    # get the last row in the existing Excel sheet
    # if it was not specified explicitly
    if startrow is None and sheet_name in writer.book.sheetnames:
        startrow = writer.book[sheet_name].max_row + bufferrow

    # truncate sheet
    if truncate_sheet and sheet_name in writer.book.sheetnames:
        # index of [sheet_name] sheet
        idx = writer.book.sheetnames.index(sheet_name)
        # remove [sheet_name]
        writer.book.remove(writer.book.worksheets[idx])
        # create an empty sheet [sheet_name] using old index
        writer.book.create_sheet(sheet_name, idx)
    
    # copy existing sheets
    writer.sheets = {ws.title:ws for ws in writer.book.worksheets}
    
    #hiding the sheet using sheet_visible flag
    if not sheet_visible:
        writer.sheets[sheet_name].sheet_state = 'hidden'

    if startrow is None:
        startrow = 0

    # write out the new sheet
    df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)

    # save the workbook
    writer.save()

def heatmap_corr(df):
    #-----------------------------------------------
    # Creates a heatmap of correlation from a dataframe
    #-----------------------------------------------
    import seaborn as sns #seaborn wrapper plotting library
    corrmat = df.corr() # build the matrix of correlation from the dataframe using pandas.corr() function
    f, ax = plt.subplots(figsize=(12, 9)) 	# set up the matplotlib figure
    sns.heatmap(corrmat, vmax=1.0, vmin=-1.0, square=True, annot = True) 	# draw the heatmap using seaborn

def heatmap_pbs(features, target, data):
    final_lst = []
    for col in features:
        pbs = stats.pointbiserialr(data[target], data[col])
        corr_value = pbs[0]
        p_val = pbs[1]
        final_lst.append([col, corr_value])
    pbs_df = pd.DataFrame(data = final_lst, columns=['feature', 'corr'])
    pbs_df.set_index(['feature'], inplace = True)
    f, ax = plt.subplots(figsize=(10, 8)) 	# set up the matplotlib figure
    sns.heatmap(pbs_df, vmax=1.0, vmin=-1.0, square=True, annot = True) 	# draw the heatmap using seaborn

def dt_counts(x):
    stri = 0
    num = 0
    nan = 0
    for i in x:
        if pd.isna(i):
            nan = nan+1
        elif type(i) == int or type(i) == float or str(i).isnumeric() :
            num = num+1
        else :
            stri = stri+1
    if x.dtype=='object': print('Dtype of col\tObject')
    if x.dtype=='int64':print('Dtype of col\tInt')
    if x.dtype=='float64':print('Dtype of col\tfloat64')
    print('str\t', stri)
    print('int\t', num)
    print('nan\t', nan)

def ch_dtype(x):
    try:
        if x.isnumeric():
            return int(x)
        else:
            return x
        
    except AttributeError:
        return x

def merge_size(left_frame, right_frame, on, how = 'inner', dropna = True):

    #size of each group based on groupby keys
    left_groups = left_frame.groupby(by = on).size()
    right_groups = right_frame.groupby(by = on).size()

    #finding set of common keys from two dfs
    left_keys = set(left_groups.index)
    right_keys = set(right_groups.index)
    intersection = right_keys & left_keys
    left_sub_right = left_keys - intersection
    right_sub_left = right_keys - intersection

    #to count nulls in each df on key column
    left_nan = left_frame[on].isnull().sum()[0]
    right_nan = right_frame[on].isnull().sum()[0]
    
    left_nan = 1 if left_nan == 0 and right_nan != 0 else left_nan
    right_nan = 1 if right_nan == 0 and left_nan != 0 else right_nan

    if how == 'inner':
        sizes = [(left_groups[group_name] * right_groups[group_name]) for group_name in intersection]
        if dropna == False:
            sizes += [left_nan * right_nan]
        
        return sum(sizes)

    elif how == 'left':
        sizes = [(left_groups[group_name] * right_groups[group_name]) for group_name in intersection]
        sizes += [left_groups[group_name] for group_name in left_sub_right]
        if dropna == False:
            sizes += [left_nan * right_nan]
        return sum(sizes)

    elif how == 'right':
        sizes = [(left_groups[group_name] * right_groups[group_name]) for group_name in intersection]
        sizes += [right_groups[group_name] for group_name in right_sub_left]
        if dropna == False:
            sizes += [left_nan * right_nan]
        return sum(sizes)

    elif how == 'outer':
        sizes = [(left_groups[group_name] * right_groups[group_name]) for group_name in intersection]
        sizes += [left_groups[group_name] for group_name in left_sub_right]
        sizes += [right_groups[group_name] for group_name in right_sub_left]
        if dropna == False:
            sizes += [left_nan * right_nan]
        return sum(sizes)